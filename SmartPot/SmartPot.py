from time import sleep
import csv
import serial
import tkinter
import datetime
import time
from datetime import date
import openpyxl
from openpyxl import Workbook
from openpyxl.chart import ScatterChart,LineChart, Reference, Series
from openpyxl.chart.axis import DateAxis
from openpyxl.chart.label import DataLabelList, DataLabel
from openpyxl.styles import Font, Color, colors
import smtplib,ssl
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.utils import formatdate
from email import encoders
import threading
from threading import Thread
def send_Not(strInput):
    msg = 'Subject: {}\n\n{}'.format('Condition', strInput)
    to = ''
    gmail_user = ''
    gmail_pwd = ''
    smtpserver = smtplib.SMTP("smtp.gmail.com",587)
    smtpserver.ehlo()
    smtpserver.starttls()
    smtpserver.ehlo() # extra characters to permit edit
    smtpserver.login(gmail_user, gmail_pwd)
    smtpserver.sendmail(gmail_user, to, msg)
    smtpserver.quit()
    print("Email sent")
class part1(Thread):
    tControl = 0
    def run(self):
        print("Thread Started")
        while(part1.tControl != -1) :
            if(part1.tControl == 1):
                send_Not("Too Hot")
                part1.tControl = 0
                sleep(2)
            if(part1.tControl == 2):
                send_Not("Too much water")
                part1.tControl = 0
                sleep(2)
        sleep(0.5)
        print("Thread Terminated")            
def send_Email():
    List = []
    count = 0
    f2 = open('dataLog.csv','rt')
    fileData = csv.reader(f2)
    for row in fileData:
            List.append(row)
            count += 1
            print(row)

    humidityList = []
    temperatureList = []
    lightList = [] # new

    count2 = 0
    for x in range(count) :
        if(List[x][1] == 'humidity') :
            humidityList.append(List[x])
            count2 +=1
        if(List[x][1] == 'temperature') :
            temperatureList.append(List[x])
        if(List[x][1] == 'light') : # new 
            lightList.append(List[x]) # new


    # open Workbook
    wb = Workbook()
    ws = wb.active

    # Format Values
    for x in range(count2) :
        humidityList[x][2] = int(humidityList[x][2])
        xDT = datetime.datetime.strptime(humidityList[x][0],'%Y-%m-%d %H:%M:%S.%f')
        humidityList[x][0] = xDT
    ws.append([])
    for x in range(count2) :
        temperatureList[x][2] = float(temperatureList[x][2])
        xDT = datetime.datetime.strptime(temperatureList[x][0],'%Y-%m-%d %H:%M:%S.%f')
        temperatureList[x][0] = xDT
    for x in range(count2) : #new start
        lightList[x][2] = int(lightList[x][2])
        xDT = datetime.datetime.strptime(lightList[x][0],'%Y-%m-%d %H:%M:%S.%f')
        lightList[x][0] = xDT # new end

    dTC = ws.cell(row = 1,column = 1,value='Date-Time')
    sTC = ws.cell(row = 1,column = 2,value='Sensor Type')
    vC = ws.cell(row = 1,column = 3,value='Value')
    uC = ws.cell(row = 1,column = 4,value='Unit')
    sC = ws.cell(row = 1,column = 5,value='Symbol')

    dTC.font = Font(color=colors.BLUE, italic=False)
    sTC.font = Font(color=colors.BLUE, italic=False)
    vC.font = Font(color=colors.BLUE, italic=False)
    uC.font = Font(color=colors.BLUE, italic=False)
    sC.font = Font(color=colors.BLUE, italic=False)
    # Insert ınto file
    print("-------------------")
    for x in range(count2) :
        print(humidityList[x])
        ws.append(humidityList[x])
    ws.append([])
    for x in range(count2) :
        print(temperatureList[x])
        ws.append(temperatureList[x])
    ws.append([]) # new start
    for x in range(count2) :
        print(lightList[x])
        ws.append(lightList[x]) # new end

    # Fromat Column
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 12

    # Set Chart
    chart = ScatterChart()
    #chart.title = "Temperature Chart"
    chart.x_axis.title = 'Date-Time'
    chart.y_axis.title = '*C'
    # axis
    tempMinRow = 2 + (count2) + 1
    tempMaxRow = 1 + (count2*2) + 1

    # x-axis
    xvalues = Reference(ws,min_col=1,min_row=tempMinRow,max_row=tempMaxRow)
    # y-axis
    yvalues = Reference(ws,min_col=3,min_row=tempMinRow,max_row=tempMaxRow)
    # Series
    series = Series(yvalues,xvalues)
    chart.series.append(series)
    # Style
    chart.style = 10
    chart.height = 10
    chart.width = 20
    #Show axis
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    # Set position of chart
    posTemp = (count2*3) + 2 + 4
    posTemp2 = "A"+str(posTemp)

    cTitle = ws.cell(row = posTemp-1,column = 1,value='Temperature Chart')
    cTitle.font = Font(color=colors.BLUE, italic=False)

    # Humidity Chart
    chart2 = ScatterChart()
    chart2.x_axis.title = 'Date-Time'
    chart2.y_axis.title = '%'
    humMinRow = 2
    humMaxRow = 1 + (count2)
    xvalues2 = Reference(ws,min_col=1,min_row=humMinRow,max_row=humMaxRow)
    yvalues2 = Reference(ws,min_col=3,min_row=humMinRow,max_row=humMaxRow)
    series2 = Series(yvalues2,xvalues2)
    chart2.series.append(series2)
    chart2.style = 10
    chart2.height = 10
    chart2.width = 20
    chart2.x_axis.delete = False
    chart2.y_axis.delete = False
    posHum = (count2*3) + 2 + 4 + 22
    posHum2 = "A"+str(posHum)
    cTitle2 = ws.cell(row = posHum-1,column = 1,value='Humidity Chart')
    cTitle2.font = Font(color=colors.BLUE, italic=False)

    # Light Chart
    chart3 = ScatterChart()
    chart3.x_axis.title = 'Date-Time'
    chart3.y_axis.title = 'nm'
    lightMinRow = 2 + (count2)*2 + 2
    lightMaxRow = 1 + (count2*3) + 2

    xvalues3 = Reference(ws,min_col=1,min_row=lightMinRow,max_row=lightMaxRow)
    yvalues3 = Reference(ws,min_col=3,min_row=lightMinRow,max_row=lightMaxRow)
    series3 = Series(yvalues3,xvalues3)
    chart3.series.append(series3)
    chart3.style = 10
    chart3.height = 10
    chart3.width = 20
    chart3.x_axis.delete = False
    chart3.y_axis.delete = False
    posLight = (count2*3) + 3 + 4 + 44
    posLight2 = "A"+str(posLight)
    cTitle3 = ws.cell(row = posLight-1,column = 1,value='Light Chart')
    cTitle3.font = Font(color=colors.BLUE, italic=False)

    # Add and finish
    ws.add_chart(chart,posTemp2)
    ws.add_chart(chart2,posHum2)
    ws.add_chart(chart3,posLight2)
    wb.save("Table.xlsx")

    
    file = 'Table.xlsx'
    msg = MIMEMultipart()
    fp = open(file, 'rb')
    part = MIMEBase('application','vnd.ms-excel')
    part.set_payload(fp.read())
    fp.close()
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', 'attachment', filename = 'Table.xlsx')
    msg.attach(part)
    msg['Subject'] = 'Pot Project'
    to = ''
    gmail_user = ''
    gmail_pwd = ''
    smtpserver = smtplib.SMTP("smtp.gmail.com",587)
    smtpserver.ehlo()
    smtpserver.starttls()
    smtpserver.ehlo() # extra characters to permit edit
    smtpserver.login(gmail_user, gmail_pwd)
    smtpserver.sendmail(gmail_user, to, msg.as_string())
    smtpserver.quit()
    print("Email Sent")

print("Hello")
class part2(Thread):
    def run(self):
        f = open('dataLog.csv','w')
        header = "date-time,sensor_type,data,unit,symbol"
        f.write(header + '\n')
        ser = serial.Serial(port = '/dev/ttyACM0',baudrate = 9600,bytesize= serial.EIGHTBITS,parity=serial.PARITY_NONE,timeout=1)
        try:
            ser.isOpen()
            print("Serial port is open")
        except:
            print("Error")
            SystemExit()
        if(ser.isOpen()) :
            try:
                countS = 0
                while(countS < 21):
                    data = ser.readline().decode("utf-8")
                    if(data[:1] == "1") : # Start of Valid
                        print(countS,") ",end = '')
                        print("Valid: ",end = '')
                        dataList = str.split(data,'/')
                        if(dataList[1] == "0") :
                                if(dataList[2] == "00") :
                                    part1.tControl = 2
                                if(dataList[2] == "01") :
                                    part1.tControl = 1
                        if(dataList[1] == "1") : # Write or Command
                            print( "Write: ",end = '')
                            logDateInfo = str(datetime.datetime.now())
                            if(dataList[2] == "00") : # Sensor type
                                print("Humidity Sensor: ",end = '')
                                sensData = data[7:]
                                print(sensData)
                                #print(countS)
                                f.write(logDateInfo +','+'humidity'+','+ dataList[3] + ',' + dataList[4] +',' + dataList[5])                  
                            if(dataList[2] == "01")  :  
                                print("Temperature Sensor: ",end = '')
                                sensData2 = data[7:]
                                print(sensData2)
                                #print(countS)
                                f.write(logDateInfo +','+'temperature'+','+ dataList[3] + ',' + dataList[4] +',' + dataList[5])
                            if(dataList[2] == "10") :
                                print("Light Sensor: ",end = '')
                                sensData3 = data[7:]
                                print(sensData3)
                                #print(countS)
                                f.write(logDateInfo +','+'light'+','+ dataList[3] + ',' + dataList[4] +',' + dataList[5])
                            countS = countS + 1
                f.close()
                part1.tControl = -1
                send_Email()            
                                        
            except Exception as e:
                print("error: " + str(e))
        else:
            print("Cannot open serial port")
t2 = part2()
t1 = part1()

t1.start()
t2.start()

        
  